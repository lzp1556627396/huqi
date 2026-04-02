from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from datetime import datetime, time

from .calculator import calculate_attendance, infer_shift_type
from .models import ShiftType
from .time_utils import format_late_early_text, normalize_text, parse_punch_datetime, parse_work_date

TOP_HEADER_ROW = 3
SUB_HEADER_ROW = 4
DATA_START_ROW = 5


class MissingColumnError(RuntimeError):
    pass


def process_workbook(
    input_file: str | Path,
    output_file: str | Path | None = None,
    sheet_name: str = "概况统计与打卡明细",
) -> Path:
    input_path = Path(input_file).expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"未找到输入文件：{input_path}")

    if output_file is None:
        output_path = input_path.with_name(f"{input_path.stem}_处理结果{input_path.suffix}")
    else:
        output_path = Path(output_file).expanduser().resolve()

    workbook = load_workbook(input_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表不存在：{sheet_name}")

    worksheet = workbook[sheet_name]
    _process_sheet(worksheet)
    workbook.save(output_path)
    return output_path


def _process_sheet(worksheet: Worksheet) -> None:
    date_col = _find_column(worksheet, "时间", None)
    rule_col = _find_column(worksheet, "考勤概况", "所属规则")
    shift_col = _find_column(worksheet, "考勤概况", "班次")
    style_source_col = _find_column(worksheet, "考勤概况", "考勤结果")

    punch_columns = [
        _find_column(worksheet, "上班1", "打卡时间"),
        _find_column(worksheet, "下班1", "打卡时间"),
        _find_column(worksheet, "上班2", "打卡时间"),
        _find_column(worksheet, "下班2", "打卡时间"),
        _find_column(worksheet, "上班3", "打卡时间"),
        _find_column(worksheet, "下班3", "打卡时间"),
    ]

    insert_at_col = worksheet.max_column + 1
    worksheet.insert_cols(insert_at_col, amount=3)

    shifted_rule_col = _shift_column_index(rule_col, insert_at_col, amount=3)
    shifted_shift_col = _shift_column_index(shift_col, insert_at_col, amount=3)
    shifted_punch_columns = [_shift_column_index(col, insert_at_col, amount=3) for col in punch_columns]

    regular_col = insert_at_col
    overtime_col = insert_at_col + 1
    late_early_col = insert_at_col + 2

    _write_custom_headers(worksheet, regular_col, overtime_col, late_early_col)
    _copy_column_style(worksheet, source_col=style_source_col, target_cols=[regular_col, overtime_col, late_early_col])

    for row in range(DATA_START_ROW, worksheet.max_row + 1):
        work_date = parse_work_date(worksheet.cell(row, date_col).value)
        if work_date is None:
            continue

        rule_text = _as_text(worksheet.cell(row, shifted_rule_col).value)
        shift_text = _as_text(worksheet.cell(row, shifted_shift_col).value)
        shift_type = infer_shift_type(shift_text) or infer_shift_type(rule_text)

        if shift_type is None:
            raw_punch_values = [worksheet.cell(row, col).value for col in shifted_punch_columns]
            shift_type = _infer_shift_from_punch_values(raw_punch_values)

        if shift_type is None:
            # Fall back to day-shift rules for records with unknown shift text.
            shift_type = ShiftType.DAY

        punch_moments = []
        for col in shifted_punch_columns:
            moment = parse_punch_datetime(worksheet.cell(row, col).value, work_date, shift_type)
            if moment is not None:
                punch_moments.append(moment)

        metrics = calculate_attendance(work_date, punch_moments, shift_type)
        late_early_text = format_late_early_text(
            metrics.late_minutes,
            metrics.early_leave_minutes,
            metrics.has_punch,
        )

        regular_cell = worksheet.cell(row, regular_col, metrics.regular_hours)
        overtime_cell = worksheet.cell(row, overtime_col, metrics.overtime_hours)
        worksheet.cell(row, late_early_col, late_early_text)

        regular_cell.number_format = "0.00"
        overtime_cell.number_format = "0.00"


def _write_custom_headers(
    worksheet: Worksheet,
    regular_col: int,
    overtime_col: int,
    late_early_col: int,
) -> None:
    worksheet.cell(TOP_HEADER_ROW, regular_col, "自定义统计")
    worksheet.cell(SUB_HEADER_ROW, regular_col, "正班出勤时长(小时)")
    worksheet.cell(SUB_HEADER_ROW, overtime_col, "加班时长(小时)")
    worksheet.cell(SUB_HEADER_ROW, late_early_col, "迟到早退时长")
    worksheet.merge_cells(
        start_row=TOP_HEADER_ROW,
        start_column=regular_col,
        end_row=TOP_HEADER_ROW,
        end_column=late_early_col,
    )


def _copy_column_style(worksheet: Worksheet, source_col: int, target_cols: list[int]) -> None:
    for row in (TOP_HEADER_ROW, SUB_HEADER_ROW):
        source_cell = worksheet.cell(row, source_col)
        for target_col in target_cols:
            _clone_style(source_cell, worksheet.cell(row, target_col))

    for row in range(DATA_START_ROW, worksheet.max_row + 1):
        source_cell = worksheet.cell(row, source_col)
        for target_col in target_cols:
            _clone_style(source_cell, worksheet.cell(row, target_col))


def _clone_style(source_cell: Any, target_cell: Any) -> None:
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.protection = copy(source_cell.protection)
    target_cell.number_format = source_cell.number_format


def _find_column(worksheet: Worksheet, top_header: str, sub_header: str | None) -> int:
    matches: list[int] = []
    last_top = None

    for col in range(1, worksheet.max_column + 1):
        raw_top = worksheet.cell(TOP_HEADER_ROW, col).value
        if raw_top is None:
            raw_top = last_top
        else:
            last_top = raw_top

        raw_sub = worksheet.cell(SUB_HEADER_ROW, col).value
        top = _as_text(raw_top)
        sub = _as_text(raw_sub) if raw_sub is not None else None

        if top != top_header:
            continue

        if sub_header is None:
            if sub in (None, ""):
                matches.append(col)
        elif sub == sub_header:
            matches.append(col)

    if not matches:
        raise MissingColumnError(f"未找到列：({top_header}, {sub_header})")
    return matches[0]


def _shift_column_index(column_index: int, insert_at_col: int, amount: int) -> int:
    if column_index >= insert_at_col:
        return column_index + amount
    return column_index


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _infer_shift_from_punch_values(values: list[Any]) -> ShiftType | None:
    parsed_times: list[time] = []
    has_next_day_hint = False

    for value in values:
        if value is None:
            continue
        if isinstance(value, datetime):
            parsed_times.append(value.time())
            continue
        if isinstance(value, time):
            parsed_times.append(value)
            continue

        text = normalize_text(value)
        if not text or text.lower() in {"--", "未打卡", "nan", "none", "nat", "null"}:
            continue
        if "次日" in text:
            has_next_day_hint = True

        cleaned = text.replace("次日", "").strip()
        parts = cleaned.split(":")
        if len(parts) != 2:
            continue

        hour_text, minute_text = parts
        if not (hour_text.isdigit() and minute_text.isdigit()):
            continue

        hour = int(hour_text)
        minute = int(minute_text)
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            parsed_times.append(time(hour=hour, minute=minute))

    if not parsed_times:
        return None

    if has_next_day_hint:
        return ShiftType.NIGHT

    if any(t.hour < 7 for t in parsed_times):
        return ShiftType.NIGHT

    first_time = parsed_times[0]
    if first_time.hour >= 16:
        return ShiftType.NIGHT

    return ShiftType.DAY
